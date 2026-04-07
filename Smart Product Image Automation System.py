import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.styles import PatternFill
from PIL import Image as PILImage

#                          DEFAULT PATHS 
DEFAULT_IMAGE_FOLDER = r"C:\Users\GOD\OneDrive\Desktop\Checking\IMAGE"
DEFAULT_OUTPUT_FOLDER = r"C:\Users\GOD\OneDrive\Desktop\Checking\OUTPUT"

def get_path(prompt, default=None, is_folder=False):
    """Helper function to get path from user"""
    while True:
        if default:
            user_input = input(f"{prompt} (Press Enter for: {default})\n> ").strip()
            path = user_input if user_input else default
        else:
            path = input(f"{prompt}\n> ").strip()
        
        if not path:
            print("❌ Path cannot be empty. Please try again.\n")
            continue
            
        path = os.path.abspath(os.path.expanduser(path))
        
        if is_folder:
            if not os.path.isdir(path):
                try:
                    os.makedirs(path, exist_ok=True)
                    print(f"✅ Created folder: {path}")
                except Exception as e:
                    print(f"⚠️ Could not create folder: {e}")
                    continue
            return path
        else:
            if not os.path.isfile(path):
                print(f"⚠️ File not found: {path}")
                print("Please enter a valid file path.\n")
                continue
            return path


print("🖼️  Smart Product Image Automation System\n")

#                        ONLY ASK FOR EXCEL FILE PATH
excel_file = get_path("Enter the path of your Excel file (.xlsx)", 
                      default=None, is_folder=False)

#                Use default Paths for Image and Output folders
image_folder = get_path("Enter the path of the IMAGE folder", 
                        default=DEFAULT_IMAGE_FOLDER, is_folder=True)

output_folder = get_path("Enter the path for OUTPUT folder", 
                         default=DEFAULT_OUTPUT_FOLDER, is_folder=True)

print("\n" + "="*70)
print("🚀 Starting processing...")
print("="*70 + "\n")

#                                    Formatting  
CELL_SIZE_PIXELS = 120          
PADDING_PIXELS = 5            

IMAGE_DISPLAY_WIDTH = CELL_SIZE_PIXELS - (2 * PADDING_PIXELS)
IMAGE_DISPLAY_HEIGHT = CELL_SIZE_PIXELS - (2 * PADDING_PIXELS)

ROW_HEIGHT_POINTS = CELL_SIZE_PIXELS
COLUMN_WIDTH = CELL_SIZE_PIXELS / 7.5

#                        Generate output filename with timestamp
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
output_file = os.path.join(output_folder, f"Image Inserted File {timestamp}.xlsx")

#                                    LOAD WORKBOOK 
wb = load_workbook(excel_file)
ws = wb.active

#                                     FIND COLUMNS 
header = {}
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    if val:
        header[str(val).strip().upper()] = col

if "ARTICLE IMAGE" not in header or "UPC" not in header:
    raise ValueError("Required columns 'ARTICLE IMAGE' or 'UPC' not found in header row!")

image_col = header["ARTICLE IMAGE"]
article_col = header["UPC"]

#                               Supported image extensions
image_extensions = [ext.lower() for ext, f in PILImage.registered_extensions().items() 
                   if f in PILImage.OPEN]

images_inserted = 0
images_not_found = 0
temp_files = []

no_fill = PatternFill(fill_type=None)

print("Processing rows...\n")

for row in range(2, ws.max_row + 1):
    article_no = ws.cell(row=row, column=article_col).value
    if not article_no:
        continue
    
    article_no = str(article_no).strip()
    image_found = False

    for ext in image_extensions:
        img_path = os.path.join(image_folder, article_no + ext)
        
        if os.path.exists(img_path):
            try:
                pil_img = PILImage.open(img_path)
                
                temp_path = os.path.join(output_folder, f"temp_{article_no}_{row}.png")
                pil_img.save(temp_path, format='PNG', quality=100)
                temp_files.append(temp_path)

                img = Image(temp_path)
                img.width = IMAGE_DISPLAY_WIDTH
                img.height = IMAGE_DISPLAY_HEIGHT

                col_letter = ws.cell(row=1, column=image_col).column_letter

                
                col_offset = pixels_to_EMU(PADDING_PIXELS)
                cell_height_emu = pixels_to_EMU(ROW_HEIGHT_POINTS)
                image_height_emu = pixels_to_EMU(IMAGE_DISPLAY_HEIGHT)
                vertical_offset = (cell_height_emu - image_height_emu) // 2

                marker = AnchorMarker(
                    col=image_col - 1,
                    colOff=col_offset,
                    row=row - 1,
                    rowOff=vertical_offset
                )

                img.anchor = OneCellAnchor(_from=marker)

                ws.add_image(img, f"{col_letter}{row}")

               
                ws.row_dimensions[row].height = ROW_HEIGHT_POINTS
                ws.column_dimensions[col_letter].width = COLUMN_WIDTH

                ws.cell(row=row, column=image_col).fill = no_fill

                images_inserted += 1
                image_found = True
                print(f"✅ Inserted: {article_no}{ext}  (Row {row})")
                break

            except Exception as e:
                print(f"⚠️ Error processing {article_no}{ext}: {e}")

    if not image_found:
        print(f"❌ Image not found for: {article_no}")
        images_not_found += 1

#                                     SAVE & CLEANUP
wb.save(output_file)

for temp in temp_files:
    try:
        if os.path.exists(temp):
            os.remove(temp)
    except:
        pass

#                                       SUMMARY 
print("\n" + "="*80)
print("🎉 PROCESS COMPLETED SUCCESSFULLY!")
print("="*80)
print(f"📁 Output File     : {output_file}")
print(f"✅ Images Inserted : {images_inserted}")
print(f"❌ Images Not Found: {images_not_found}")
print("="*80)